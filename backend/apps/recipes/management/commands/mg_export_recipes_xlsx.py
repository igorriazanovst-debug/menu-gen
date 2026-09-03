"""MG_RECIPEXLSX: выгрузка рецептов в таблицу для работы редактора.

Нужна, чтобы обходить импортированные рецепты списком и доснимать к ним фото:
у say7-выгрузки фотографий нет, поэтому рецепты лежат неопубликованными.

Пять колонок: id, название, шаги одной ячейкой с нумерацией, состав одной
ячейкой, ссылка.

Про ссылку. У карточки рецепта нет собственного адреса во фронте: роут только
`/recipes` (список), параметр `?id=` страница не читает. Поэтому по умолчанию
ставится ссылка в админку — она открывает конкретный рецепт и ведёт туда, где
фото и загружают. `--link app` даёт адрес веб-приложения, но он откроет список,
а не карточку; оставлен на случай, если во фронте появится нормальный роут.

    python manage.py mg_export_recipes_xlsx
    python manage.py mg_export_recipes_xlsx --out /app/media/say7.xlsx
    python manage.py mg_export_recipes_xlsx --legacy-prefix tg: --link app
    python manage.py mg_export_recipes_xlsx --csv        # если openpyxl нет

MG_UNUSABLE: та же выгрузка умеет отбирать негодные к публикации — те, у
которых нет веса порции или калорий на порцию. Такой рецепт можно снять и
опубликовать, но генератор меню его никогда не возьмёт: без веса порции он не
проходит ни коридор калорий, ни «Тарелку». Фотографировать их — выброшенная
работа, и список нужен как раз чтобы этого не делать.

    python manage.py mg_export_recipes_xlsx --unpublished --missing-portion
    python manage.py mg_export_recipes_xlsx --unpublished --missing-portion --no-photo
    python manage.py mg_export_recipes_xlsx --unpublished --missing-portion --dish soup

--no-photo отделяет безнадёжных от поправимых: рецепт без веса порции, но с
фотографией, ещё можно довести руками (проставить вес — и он годен). Тот же
рецепт без фотографии требует и съёмки, и ручной правки, то есть двойной
работы на каждую строку.
"""

import csv
import os

from django.core.management.base import BaseCommand, CommandError

from apps.recipes.models import Recipe

HEADERS = ["ID", "Название", "Шаги", "Состав", "Ссылка"]

# MG_UNUSABLE: колонки, объясняющие, ПОЧЕМУ рецепт негоден. Без них таблица
# says «вот эти плохие» и не даёт проверить утверждение.
DIAG_HEADERS = ["Тип блюда", "Порций", "Вес порции", "Ккал/порция", "Ккал/100 г", "Фото"]


def steps_text(recipe):
    """Шаги одной ячейкой, с номерами: «1. Промыть гречку.»"""
    rows = []
    for n, step in enumerate(recipe.steps or [], 1):
        if isinstance(step, dict):
            text = (step.get("text") or "").strip()
            order = step.get("order") or n
        else:
            text, order = str(step).strip(), n
        if text:
            rows.append("%s. %s" % (order, text))
    return "\n".join(rows)


def ingredients_text(recipe):
    """Состав одной ячейкой: «Гречка — 1 стакан (200 г)».

    Количество и вес показываются, только если они есть: пустые скобки и
    висящее тире читаются хуже, чем просто название.
    """
    rows = []
    for ing in recipe.ingredients or []:
        if not isinstance(ing, dict):
            rows.append(str(ing).strip())
            continue
        name = (ing.get("name") or "").strip()
        if not name:
            continue
        сколько = " ".join(x for x in [str(ing.get("quantity") or "").strip(), (ing.get("unit") or "").strip()] if x)
        граммы = ing.get("grams")
        часть = name
        if сколько:
            часть += " — " + сколько
        if граммы:
            try:
                часть += " (%g г)" % float(граммы)
            except (TypeError, ValueError):
                pass
        rows.append(часть)
    return "\n".join(rows)


class Command(BaseCommand):
    help = "MG_RECIPEXLSX: выгрузить рецепты в xlsx (id, название, шаги, состав, ссылка)."

    def add_arguments(self, parser):
        parser.add_argument(
            "--legacy-prefix",
            default="say7:",
            help="Какие рецепты выгружать по началу legacy_id. Пусто — все.",
        )
        parser.add_argument("--out", default="", help="Путь к файлу. По умолчанию — в MEDIA_ROOT.")
        parser.add_argument(
            "--link",
            choices=["admin", "app"],
            default="admin",
            help="admin — ссылка в админку (по умолчанию); app — адрес веб-приложения.",
        )
        parser.add_argument("--base-url", default="https://menugen.ru", help="Адрес сайта без завершающего слэша.")
        parser.add_argument("--csv", action="store_true", help="Писать CSV вместо xlsx (без openpyxl).")
        parser.add_argument("--unpublished", action="store_true", help="Только неопубликованные.")
        parser.add_argument(
            "--missing-portion",
            action="store_true",
            help="Только без веса порции или без калорий на порцию (негодные к публикации).",
        )
        parser.add_argument("--dish", default="", help="Ограничить типом блюда (soup, snack, dessert, ...).")
        parser.add_argument(
            "--exclude-dish",
            default="",
            help="Исключить типы блюд через запятую, напр. soup или soup,salad.",
        )
        parser.add_argument("--no-photo", action="store_true", help="Только без фотографии.")
        parser.add_argument("--with-photo", action="store_true", help="Только с фотографией.")

    def out_path(self, opts):
        from django.conf import settings

        if opts["out"]:
            return opts["out"]
        name = "recipes_export.csv" if opts["csv"] else "recipes_export.xlsx"
        return os.path.join(settings.MEDIA_ROOT, name)

    def link_for(self, recipe, opts):
        base = opts["base_url"].rstrip("/")
        if opts["link"] == "admin":
            return "%s/admin/recipes/recipe/%s/change/" % (base, recipe.id)
        return "%s/recipes?id=%s" % (base, recipe.id)

    def headers(self, opts):
        return HEADERS + DIAG_HEADERS if opts["missing_portion"] else HEADERS

    def queryset(self, opts):
        from django.db.models import Q

        qs = Recipe.objects.all().order_by("id")
        prefix = opts["legacy_prefix"]
        if prefix:
            qs = qs.filter(legacy_id__startswith=prefix)
        if opts["unpublished"]:
            qs = qs.filter(is_published=False)
        if opts["dish"]:
            qs = qs.filter(dish_type=opts["dish"])
        if opts["exclude_dish"]:
            qs = qs.exclude(dish_type__in=[d.strip() for d in opts["exclude_dish"].split(",") if d.strip()])
        if opts["missing_portion"]:
            qs = qs.filter(Q(portion_g__isnull=True) | Q(kcal__isnull=True))
        # Пустая строка и NULL — оба означают «фото нет»: часть импортов пишет
        # одно, часть другое, и проверка по одному из них молча теряет половину.
        if opts["no_photo"]:
            qs = qs.filter(Q(image_url__isnull=True) | Q(image_url=""))
        if opts["with_photo"]:
            qs = qs.exclude(Q(image_url__isnull=True) | Q(image_url=""))
        return qs

    def rows(self, opts):
        for recipe in self.queryset(opts).iterator():
            row = [
                recipe.id,
                recipe.title or "",
                steps_text(recipe),
                ingredients_text(recipe),
                self.link_for(recipe, opts),
            ]
            if opts["missing_portion"]:
                row += [
                    recipe.dish_type or "",
                    recipe.servings if recipe.servings else "",
                    recipe.portion_g if recipe.portion_g else "",
                    recipe.kcal if recipe.kcal is not None else "",
                    recipe.kcal_per_100g if recipe.kcal_per_100g is not None else "",
                    "есть" if (recipe.image_url or "").strip() else "нет",
                ]
            yield row

    def handle(self, *args, **opts):
        path = self.out_path(opts)
        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)

        if opts["csv"]:
            n = self.write_csv(path, opts)
        else:
            n = self.write_xlsx(path, opts)

        self.stdout.write(self.style.SUCCESS("Записано строк: %d" % n))
        self.stdout.write("Файл: %s" % path)

        # MG_UNUSABLE: разбивка по типам блюд — по ней видно, какие роли
        # опустеют после удаления, ещё до того как что-то удалено.
        if opts["missing_portion"]:
            from collections import Counter

            by_dish = Counter(self.queryset(opts).values_list("dish_type", flat=True))
            self.stdout.write("\nПо типам блюд:")
            for dish, count in by_dish.most_common():
                self.stdout.write("    %-16s %4d" % (dish or "(не задан)", count))

    def write_csv(self, path, opts):
        # utf-8-sig: без BOM Excel открывает кириллицу кракозябрами.
        with open(path, "w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.writer(fh, delimiter=";")
            writer.writerow(self.headers(opts))
            n = 0
            for row in self.rows(opts):
                writer.writerow(row)
                n += 1
        return n

    def write_xlsx(self, path, opts):
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font
        except ImportError:
            raise CommandError(
                "openpyxl не установлен в backend-контейнере.\n"
                "Поставить: docker compose exec -T backend pip install openpyxl\n"
                "Либо выгрузить в CSV: manage.py mg_export_recipes_xlsx --csv"
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Рецепты"
        ws.append(self.headers(opts))
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"

        n = 0
        for row in self.rows(opts):
            ws.append(row)
            n += 1

        # Шаги и состав — многострочные: без переноса ячейка покажет одну строку.
        for column, width in zip("ABCDE", [8, 40, 90, 50, 60]):
            ws.column_dimensions[column].width = width
        wrap = Alignment(wrap_text=True, vertical="top")
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=4):
            for cell in row:
                cell.alignment = wrap

        wb.save(path)
        return n
