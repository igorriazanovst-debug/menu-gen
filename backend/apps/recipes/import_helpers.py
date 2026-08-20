# MG_IMPORT_TOOL_V1 — shared parse/compute helpers (used by management command + admin tool)
from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation

DISH_TYPES = {"soup", "main", "salad", "side", "dessert", "drink", "bakery", "sauce", "snack", "breakfast_dish"}
FOOD_GROUPS = {"grain", "protein", "vegetable", "fruit", "dairy", "oil", "other"}
PROTEIN_TYPES = {"animal", "plant", "mixed"}
GRAIN_TYPES = {"whole", "refined"}
COOKING_METHODS = {"boiled", "baked", "fried", "grilled", "raw", "stewed", "steamed", "microwave"}
SOURCES = {"own", "import", "user", "parsed"}
KCAL_100G_MIN = 40
KCAL_100G_MAX = 550
MEAL_OK = {"breakfast", "lunch", "dinner", "snack"}
ALLERG_OK = {"nuts", "eggs", "fish", "shellfish", "milk", "gluten", "soy", "peanuts", "sesame"}
EXAMPLE_TITLES = {"Борщ классический", "Овсяная каша на молоке"}

COLS = [
    "row_id",
    "title",
    "description",
    "ingredients",
    "steps",
    "cook_time_min",
    "difficulty",
    "image_url",
    "video_url",
    "portion_g",
    "kcal_per_100g",
    "proteins_per_100g",
    "fats_per_100g",
    "carbs_per_100g",
    "sugars_per_100g",
    "dish_type",
    "meal_category",
    "food_group",
    "protein_type",
    "grain_type",
    "is_red_meat",
    "is_fatty_fish",
    "has_added_sugar",
    "cooking_method",
    "oil_tsp",
    "is_vegan",
    "is_vegetarian",
    "is_gluten_free",
    "is_lactose_free",
    "allergens",
    "source",
    "country",
]


def s(v):
    if v is None:
        return ""
    return str(v).strip()


def num(v):
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return float(str(v).replace(",", "."))
    except (ValueError, TypeError):
        return None


def dec(v):
    n = num(v)
    if n is None:
        return None
    try:
        return Decimal(str(round(n, 1)))
    except (InvalidOperation, ValueError):
        return None


def bool_val(v):
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    return str(v).strip().upper() in {"TRUE", "ДА", "YES", "1"}


def parse_csv(v, allowed=None):
    out = []
    for part in s(v).split(","):
        p = part.strip()
        if not p:
            continue
        if allowed is None or p in allowed:
            out.append(p)
    return out


def parse_ingredients(v):
    items = []
    for line in s(v).splitlines():
        line = line.strip()
        if not line:
            continue
        parts = [p.strip() for p in line.split("|")]
        while len(parts) < 4:
            parts.append("")
        name, qty, unit, grams = parts[0], parts[1], parts[2], parts[3]
        if not name:
            continue
        items.append({"name": name, "quantity": qty, "unit": unit, "grams": num(grams) or 0})
    return items


def parse_steps(v):
    steps = []
    for i, line in enumerate(s(v).splitlines(), start=1):
        t = line.strip()
        if not t:
            continue
        t = re.sub(r"^\d+[\.\)]\s*", "", t)
        steps.append({"order": i, "text": t})
    return steps


def read_xlsx_rows(path, sheet="Recipes"):
    """Читает xlsx, возвращает list[dict] или бросает ValueError."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise ValueError("openpyxl не установлен.")
    wb = load_workbook(path, data_only=True)
    if sheet not in wb.sheetnames:
        raise ValueError(f"Лист {sheet!r} не найден. Есть: {wb.sheetnames}")
    ws = wb[sheet]
    hdr = [s(ws.cell(1, j).value) for j in range(1, ws.max_column + 1)]
    idx = {name: j for j, name in enumerate(hdr, start=1)}
    missing = [c for c in COLS if c not in idx]
    if missing:
        raise ValueError(f"В шаблоне нет колонок: {missing}")
    rows = []
    for r in range(3, ws.max_row + 1):
        rec = {c: ws.cell(r, idx[c]).value for c in COLS}
        if all(s(rec[c]) == "" for c in COLS):
            continue
        rec["_row"] = r
        rows.append(rec)
    return rows


def parse_row(rec):
    """Парсит одну строку xlsx в dict. Возвращает (data_dict, errors_list)."""
    row = rec["_row"]
    title = s(rec["title"])
    if title in EXAMPLE_TITLES:
        return None, [f"строка {row}: пример"]
    tag = f"строка {row} ({title!r})"
    errors = []

    # skip fully empty template rows (only row_id filled, no content)
    if s(rec["title"]) == "" and s(rec["ingredients"]) == "" and s(rec["steps"]) == "":
        return None, []

    req = ["title", "ingredients", "steps", "portion_g", "dish_type", "meal_category", "food_group"]
    for f in req:
        if s(rec[f]) == "":
            errors.append(f"{tag}: пусто обязательное '{f}'")

    dish_type = s(rec["dish_type"]) or None
    if dish_type and dish_type not in DISH_TYPES:
        errors.append(f"{tag}: dish_type={dish_type!r} вне списка")
    fg = s(rec["food_group"]) or None
    if fg and fg not in FOOD_GROUPS:
        errors.append(f"{tag}: food_group={fg!r} вне списка")
    pt = s(rec["protein_type"]) or None
    if pt == "dairy":
        pt = None
    if pt and pt not in PROTEIN_TYPES:
        errors.append(f"{tag}: protein_type={pt!r} вне списка")
    gt = s(rec["grain_type"]) or None
    if gt and gt not in GRAIN_TYPES:
        errors.append(f"{tag}: grain_type={gt!r} вне списка")
    cm = s(rec["cooking_method"]) or None
    if cm and cm not in COOKING_METHODS:
        errors.append(f"{tag}: cooking_method={cm!r} вне списка")
    src = s(rec["source"]) or "own"
    if src not in SOURCES:
        errors.append(f"{tag}: source={src!r} вне списка")

    meal_category = parse_csv(rec["meal_category"], MEAL_OK)
    allergens = parse_csv(rec["allergens"])
    ingredients = parse_ingredients(rec["ingredients"])
    steps_parsed = parse_steps(rec["steps"])
    if not ingredients:
        errors.append(f"{tag}: ingredients пусты после разбора")
    if not steps_parsed:
        errors.append(f"{tag}: steps пусты после разбора")

    data = {
        "tag": tag,
        "row": row,
        "title": title,
        "description": s(rec["description"]),
        "ingredients": ingredients,
        "steps": steps_parsed,
        "cook_time_min": int(num(rec["cook_time_min"])) if num(rec["cook_time_min"]) else None,
        "image_url": s(rec["image_url"]) or None,
        "video_url": s(rec["video_url"]) or None,
        "portion_g": int(num(rec["portion_g"])) if num(rec["portion_g"]) else None,
        "kcal_per_100g": dec(rec["kcal_per_100g"]),
        "proteins_per_100g": dec(rec["proteins_per_100g"]),
        "fats_per_100g": dec(rec["fats_per_100g"]),
        "carbs_per_100g": dec(rec["carbs_per_100g"]),
        "sugars_per_100g": dec(rec["sugars_per_100g"]),
        "dish_type": dish_type,
        "food_group": fg,
        "protein_type": pt,
        "grain_type": gt,
        "is_red_meat": bool_val(rec["is_red_meat"]),
        "is_fatty_fish": bool_val(rec["is_fatty_fish"]),
        "has_added_sugar": bool_val(rec["has_added_sugar"]),
        "cooking_method": cm,
        "oil_tsp": dec(rec["oil_tsp"]),
        "is_vegan": bool_val(rec["is_vegan"]),
        "is_vegetarian": bool_val(rec["is_vegetarian"]),
        "is_gluten_free": bool_val(rec["is_gluten_free"]),
        "is_lactose_free": bool_val(rec["is_lactose_free"]),
        "allergens": allergens,
        "source": src,
        "country": s(rec["country"]) or None,
        "meal_category": meal_category,
    }
    return data, errors


def compute_kbju_100g(ingredients, kcal_anchor, use_ai=True):
    """Считает КБЖУ на 100г из ингредиентов. Возвращает (dict|None, unresolved_list)."""
    from apps.fridge.models import Product

    try:
        from apps.fridge.services import gpt_fill_nutrition
    except Exception:
        gpt_fill_nutrition = None

    tot_g = 0.0
    agg = {"proteins": 0.0, "fats": 0.0, "carbs": 0.0, "sugars": 0.0, "kcal": 0.0}
    unresolved = []
    for ing in ingredients:
        g = float(ing.get("grams") or 0)
        if g <= 0:
            continue
        name = ing["name"]
        cals100, nut100 = None, {}
        # MG_PRODFAMILY: КБЖУ берём из каталога, не из продуктов семей.
        from apps.fridge.visibility import catalog_q

        catalog = Product.objects.filter(catalog_q())
        p = catalog.filter(name__iexact=name).first() or catalog.filter(name__icontains=name[:20]).first()
        if p and p.calories_per_100g is not None:
            cals100 = float(p.calories_per_100g)
            nut100 = dict(p.nutrition or {})
        elif use_ai and gpt_fill_nutrition:
            cals100, nut100 = gpt_fill_nutrition(name)
        if cals100 is None and not nut100:
            unresolved.append(name)
            continue
        tot_g += g
        f = g / 100.0
        agg["kcal"] += (cals100 or 0) * f
        for k in ("proteins", "fats", "carbs", "sugars"):
            agg[k] += float(nut100.get(k) or 0) * f

    if tot_g <= 0:
        return None, unresolved

    per100 = {k: (agg[k] / tot_g * 100.0) for k in agg}
    if kcal_anchor and per100["kcal"] > 0:
        scale = max(0.5, min(2.0, float(kcal_anchor) / per100["kcal"]))
        for k in ("proteins", "fats", "carbs", "sugars"):
            per100[k] *= scale

    return {
        "kcal": round(per100["kcal"], 1),
        "proteins": round(per100["proteins"], 1),
        "fats": round(per100["fats"], 1),
        "carbs": round(per100["carbs"], 1),
        "sugars": round(per100["sugars"], 1),
    }, unresolved


def enrich_kbju(data_list, use_ai=True):
    """
    Дозаполняет КБЖУ для каждого item в data_list на месте.
    Возвращает список предупреждений (строки).
    """
    warnings = []
    for d in data_list:
        need = d["proteins_per_100g"] is None or d["fats_per_100g"] is None or d["carbs_per_100g"] is None
        if not need:
            continue
        kbju, unresolved = compute_kbju_100g(d["ingredients"], d["kcal_per_100g"], use_ai)
        if kbju:
            if d["kcal_per_100g"] is None and kbju.get("kcal"):
                d["kcal_per_100g"] = Decimal(str(kbju["kcal"]))
                kc = float(kbju["kcal"])
                if kc < KCAL_100G_MIN or kc > KCAL_100G_MAX:
                    warnings.append(f"{d['tag']}: ⚠ kcal/100г={kc} вне нормы ({KCAL_100G_MIN}..{KCAL_100G_MAX})")
            if d["proteins_per_100g"] is None:
                d["proteins_per_100g"] = Decimal(str(kbju["proteins"]))
            if d["fats_per_100g"] is None:
                d["fats_per_100g"] = Decimal(str(kbju["fats"]))
            if d["carbs_per_100g"] is None:
                d["carbs_per_100g"] = Decimal(str(kbju["carbs"]))
            if d["sugars_per_100g"] is None and kbju.get("sugars"):
                d["sugars_per_100g"] = Decimal(str(kbju["sugars"]))
            if unresolved:
                warnings.append(f"{d['tag']}: не найдены ингредиенты: {unresolved}")
        else:
            warnings.append(f"{d['tag']}: ⚠ КБЖУ посчитать не удалось")
    return warnings


def save_recipes(data_list):
    """Сохраняет список рецептов в БД. Возвращает кол-во созданных."""
    from decimal import Decimal as _D

    from django.db import transaction

    from apps.recipes.models import Recipe

    created = 0
    with transaction.atomic():
        for d in data_list:
            portion = d["portion_g"] or 0
            k100 = d["kcal_per_100g"]

            def pp(x):
                if x is None or not portion:
                    return None
                return _D(str(round(float(x) * portion / 100.0, 1)))

            nutrition_json = {}
            for key, fld in [
                ("calories", "kcal_per_100g"),
                ("proteins", "proteins_per_100g"),
                ("fats", "fats_per_100g"),
                ("carbs", "carbs_per_100g"),
                ("sugars", "sugars_per_100g"),
            ]:
                val = d.get(fld) if fld != "kcal_per_100g" else k100
                if val is not None:
                    nutrition_json[key] = float(val)

            Recipe.objects.create(
                title=d["title"],
                ingredients=d["ingredients"],
                steps=d["steps"],
                image_url=d["image_url"],
                video_url=d["video_url"],
                country=d["country"],
                is_published=True,
                is_custom=False,
                dish_type=d["dish_type"],
                portion_g=d["portion_g"],
                kcal_per_100g=k100,
                proteins_per_100g=d["proteins_per_100g"],
                fats_per_100g=d["fats_per_100g"],
                carbs_per_100g=d["carbs_per_100g"],
                sugars_per_100g=d["sugars_per_100g"],
                cook_time_min=d["cook_time_min"],
                is_vegan=d["is_vegan"],
                is_vegetarian=d["is_vegetarian"],
                is_gluten_free=d["is_gluten_free"],
                is_lactose_free=d["is_lactose_free"],
                allergens=d["allergens"],
                source=d["source"],
                food_group=d["food_group"],
                protein_type=d["protein_type"],
                grain_type=d["grain_type"],
                is_red_meat=d["is_red_meat"],
                is_fatty_fish=d["is_fatty_fish"],
                has_added_sugar=d["has_added_sugar"],
                cooking_method=d["cooking_method"],
                oil_tsp=d["oil_tsp"],
                suitable_for=d["meal_category"],
                cook_time=(f"{d['cook_time_min']} мин" if d["cook_time_min"] else None),
                nutrition=nutrition_json,
                kcal=pp(k100),
                proteins=pp(d["proteins_per_100g"]),
                fats=pp(d["fats_per_100g"]),
                carbs=pp(d["carbs_per_100g"]),
                servings=1,
                servings_normalized=1,
                serving_size_label=(f"1 порция / {portion} г" if portion else None),
            )
            created += 1
    return created
