#!/usr/bin/env python3
"""MG-204: пометить как ✅ Done в MenuGen_Backlog.xlsx"""
import sys
from pathlib import Path
from openpyxl import load_workbook

XLSX = "/opt/menugen/MenuGen_Backlog.xlsx"

if not Path(XLSX).exists():
    print(f"NO {XLSX}")
    sys.exit(1)

wb = load_workbook(XLSX)
ws = wb["Backlog"]

# header: ID Приоритет Категория Спринт Задача Описание ...
# найдём строку с MG-204 и поменяем "Задача" (col E) на "✅ ..."
# и добавим в "Описание" пометку "(web сделано, mobile отложен)"

ID_COL = 1
TITLE_COL = 5
DESC_COL = 6

found = False
for row in ws.iter_rows(min_row=2, values_only=False):
    if row[ID_COL - 1].value == "MG-204":
        old_title = row[TITLE_COL - 1].value or ""
        if not old_title.startswith("✅"):
            row[TITLE_COL - 1].value = "✅ (web) " + old_title
        old_desc = row[DESC_COL - 1].value or ""
        marker = "[MG-204 web сделано — типы FamilyMember расширены, FamilyMemberEditModal, DayNutritionSummary в MenuPage. Mobile отложен.]"
        if marker not in old_desc:
            row[DESC_COL - 1].value = old_desc + " " + marker
        found = True
        print(f"  MG-204 → '{row[TITLE_COL - 1].value}'")
        break

if not found:
    print("MG-204 не найден")
    sys.exit(1)

wb.save(XLSX)
print(f"Saved {XLSX}")
