"""
MG-205 финал: обновление MenuGen_Backlog.xlsx
  1) MG-205: статус → ✅ Done, добавить ссылки на скрипты
  2) Добавить новую карточку MG-205-UI (P2, 4ч)

Запуск (из контейнера, БД-агностичен — работает только с .xlsx):
  docker compose -f /opt/menugen/docker-compose.yml exec -T backend \
    python /app/scripts/mg_205_update_backlog.py

Идемпотентен:
  - MG-205: меняет статус только если он != 'Done'
  - MG-205-UI: добавляет, только если строка отсутствует
Бэкап .xlsx с TS-суффиксом сохраняется рядом.
"""
from __future__ import annotations
import shutil
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

XLSX_PATH = Path("/app/data/MenuGen_Backlog.xlsx")
SHEET = "Backlog"

# Альтернативные пути на случай иной структуры
CANDIDATES = [
    XLSX_PATH,
    Path("/opt/menugen/MenuGen_Backlog.xlsx"),
    Path("/app/MenuGen_Backlog.xlsx"),
    Path("/app/backend/MenuGen_Backlog.xlsx"),
    Path("/opt/menugen/backend/MenuGen_Backlog.xlsx"),
]


def find_xlsx() -> Path:
    for p in CANDIDATES:
        if p.exists():
            return p
    raise FileNotFoundError(f"MenuGen_Backlog.xlsx не найден ни в одном из: {CANDIDATES}")


def main():
    src = find_xlsx()
    print(f"[i] using {src}")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    bak = src.with_suffix(src.suffix + f".bak_mg205_final_{ts}")
    shutil.copy2(src, bak)
    print(f"[backup] {bak}")

    wb = load_workbook(src)
    if SHEET not in wb.sheetnames:
        raise RuntimeError(f"лист '{SHEET}' не найден; листы: {wb.sheetnames}")
    ws = wb[SHEET]

    # Определяем колонки по заголовкам (1-я строка)
    headers = {}
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[str(cell.value).strip()] = col_idx
    print(f"[i] headers: {list(headers.keys())}")

    # Пытаемся найти стандартные имена
    def col(*names):
        for n in names:
            if n in headers:
                return headers[n]
        raise KeyError(f"ни одного из заголовков {names} нет в листе")

    col_id = col("ID", "id", "Ключ", "Key")
    col_status = col("Статус", "Status", "статус")
    col_desc = col("Описание", "Description", "описание", "Title", "Заголовок")
    col_prio = None
    for n in ("Приоритет", "Priority", "приоритет", "P"):
        if n in headers:
            col_prio = headers[n]; break
    col_hours = None
    for n in ("Часов", "Hours", "Estimate", "Оценка", "часов"):
        if n in headers:
            col_hours = headers[n]; break

    # Обходим строки, ищем MG-205 и проверяем наличие MG-205-UI
    mg205_row = None
    mg205_ui_exists = False
    last_data_row = 1
    for row_idx in range(2, ws.max_row + 1):
        v = ws.cell(row=row_idx, column=col_id).value
        if v is None:
            continue
        last_data_row = row_idx
        v_str = str(v).strip()
        if v_str == "MG-205":
            mg205_row = row_idx
        if v_str == "MG-205-UI":
            mg205_ui_exists = True

    # 1) MG-205 → Done
    if mg205_row is None:
        print("[!] MG-205 не найдена — пропускаю обновление статуса")
    else:
        cur_status = ws.cell(row=mg205_row, column=col_status).value
        if str(cur_status).strip() in ("Done", "✅", "✅ Done", "done"):
            print(f"[skip] MG-205 уже в статусе '{cur_status}'")
        else:
            ws.cell(row=mg205_row, column=col_status, value="✅ Done")
            ws.cell(row=mg205_row, column=col_status).font = Font(bold=True, color="006100")
            print(f"[update] MG-205 row={mg205_row}: '{cur_status}' → '✅ Done'")

    # 2) MG-205-UI добавить
    if mg205_ui_exists:
        print("[skip] MG-205-UI уже присутствует")
    else:
        new_row = last_data_row + 1
        # Заполняем основные колонки
        ws.cell(row=new_row, column=col_id, value="MG-205-UI")
        ws.cell(row=new_row, column=col_status, value="⏳ Todo")
        if col_prio:
            ws.cell(row=new_row, column=col_prio, value="P2")
        if col_hours:
            ws.cell(row=new_row, column=col_hours, value=4)
        ws.cell(
            row=new_row,
            column=col_desc,
            value=(
                "UI: бейджи источника правок КБЖУ (auto/user/specialist N) + "
                "кнопка 'Сбросить к авто' + история изменений в модалке. "
                "Зависит от: MG-205 (✅), MG-204."
            ),
        )
        # Подсветка зелёным новой строки (как новой задачи)
        fill = PatternFill(start_color="FFFFEB99", end_color="FFFFEB99", fill_type="solid")
        for c in range(1, ws.max_column + 1):
            ws.cell(row=new_row, column=c).fill = fill
        print(f"[add]  MG-205-UI вставлена в row={new_row}")

    wb.save(src)
    print(f"[done] saved → {src}")
    print(f"[rollback] cp {bak} {src}")


if __name__ == "__main__":
    main()
